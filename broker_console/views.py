import os
import uuid
import csv
import io
from datetime import date, datetime
from django.conf import settings
from django.core.files.storage import default_storage
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import HttpResponse
import pandas as pd
from .models import (
    Broker, Employer, Carrier, Plan, PlanPremium, 
    EmployerOffering, CarrierCsvTemplate, ExportJob,
    Employee, Dependent, EnrollmentPeriod, EmployeeEnrollment,
    PlanEnrollment, EnrollmentEvent, EmployeeFormSubmission, EmployeePortalUser
)
from .serializers import (
    BrokerSerializer, EmployerSerializer, EmployerDetailSerializer,
    CarrierSerializer, PlanSerializer, PlanDetailSerializer,
    PlanPremiumSerializer, EmployerOfferingSerializer,
    CarrierCsvTemplateSerializer, ExportJobSerializer,
    EmployeeSerializer, EmployeeDetailSerializer, DependentSerializer,
    EnrollmentPeriodSerializer, EnrollmentPeriodDetailSerializer,
    EmployeeEnrollmentSerializer, EmployeeEnrollmentDetailSerializer,
    EmployeeEnrollmentSummarySerializer, PlanEnrollmentSerializer,
    EnrollmentEventSerializer, EmployeeFormSubmissionSerializer,
    EmployeeFormSubmissionListSerializer, EmployeePortalUserSerializer,
    EmployeePortalLoginSerializer, EmployeePortalRegisterSerializer
)

class BrokerViewSet(viewsets.ModelViewSet):
    queryset = Broker.objects.all()
    serializer_class = BrokerSerializer

class CarrierViewSet(viewsets.ModelViewSet):
    queryset = Carrier.objects.all()
    serializer_class = CarrierSerializer

class PlanViewSet(viewsets.ModelViewSet):
    queryset = Plan.objects.all()
    serializer_class = PlanSerializer
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return PlanDetailSerializer
        return PlanSerializer
    
    @action(detail=False, methods=['get'])
    def by_carrier(self, request):
        carrier_id = request.query_params.get('carrier_id')
        if carrier_id:
            plans = Plan.objects.filter(carrier_id=carrier_id, is_active=True)
            serializer = self.get_serializer(plans, many=True)
            return Response(serializer.data)
        return Response({'error': 'carrier_id parameter required'}, status=400)

class EmployerViewSet(viewsets.ModelViewSet):
    queryset = Employer.objects.all()
    serializer_class = EmployerSerializer
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return EmployerDetailSerializer
        return EmployerSerializer
    
    @action(detail=False, methods=['get'])
    def by_broker(self, request):
        broker_id = request.query_params.get('broker_id')
        if broker_id:
            employers = Employer.objects.filter(broker_id=broker_id)
            serializer = self.get_serializer(employers, many=True)
            return Response(serializer.data)
        return Response({'error': 'broker_id parameter required'}, status=400)
    
    @action(detail=True, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_import_employees(self, request, pk=None):
        """Bulk import employees from CSV/Excel file"""
        employer = self.get_object()
        
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        file = request.FILES['file']
        file_extension = file.name.lower().split('.')[-1]
        
        if file_extension not in ['csv', 'xlsx', 'xls']:
            return Response({
                'error': 'Unsupported file format. Please upload CSV or Excel file.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Read file content
            if file_extension == 'csv':
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
            
            # Validate required columns
            required_columns = [
                'employee_id', 'first_name', 'last_name', 'email',
                'date_of_birth', 'gender', 'hire_date'
            ]
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return Response({
                    'error': f'Missing required columns: {", ".join(missing_columns)}',
                    'required_columns': required_columns
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Process employees
            employees_created = 0
            employees_updated = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Parse dates
                    date_of_birth = pd.to_datetime(row['date_of_birth']).date()
                    hire_date = pd.to_datetime(row['hire_date']).date()
                    
                    # Prepare employee data
                    employee_data = {
                        'employer': employer,
                        'employee_id': str(row['employee_id']).strip(),
                        'first_name': str(row['first_name']).strip(),
                        'last_name': str(row['last_name']).strip(),
                        'email': str(row['email']).strip(),
                        'date_of_birth': date_of_birth,
                        'gender': str(row.get('gender', 'M')).upper()[:1],
                        'hire_date': hire_date,
                        'middle_initial': str(row.get('middle_initial', '')).strip()[:1],
                        'ssn': str(row.get('ssn', '')).strip(),
                        'phone': str(row.get('phone', '')).strip(),
                        'address_line1': str(row.get('address_line1', '')).strip(),
                        'address_line2': str(row.get('address_line2', '')).strip(),
                        'city': str(row.get('city', '')).strip(),
                        'state': str(row.get('state', '')).strip()[:2],
                        'zip_code': str(row.get('zip_code', '')).strip(),
                        'job_title': str(row.get('job_title', '')).strip(),
                        'department': str(row.get('department', '')).strip(),
                        'salary': float(row.get('salary', 0)) if pd.notna(row.get('salary')) else 0,
                        'hours_per_week': float(row.get('hours_per_week', 40)) if pd.notna(row.get('hours_per_week')) else 40,
                        'employment_status': str(row.get('employment_status', 'active')).lower(),
                        'marital_status': str(row.get('marital_status', 'single')).lower(),
                        'medical_coverage_tier': str(row.get('medical_coverage_tier', '')).strip(),
                        'dental_coverage_tier': str(row.get('dental_coverage_tier', '')).strip(),
                        'vision_coverage_tier': str(row.get('vision_coverage_tier', '')).strip(),
                    }
                    
                    # Create or update employee
                    employee, created = Employee.objects.update_or_create(
                        employer=employer,
                        employee_id=employee_data['employee_id'],
                        defaults=employee_data
                    )
                    
                    if created:
                        employees_created += 1
                    else:
                        employees_updated += 1
                        
                except Exception as e:
                    errors.append({
                        'row': index + 2,  # +2 because pandas is 0-indexed and we skip header
                        'employee_id': str(row.get('employee_id', '')),
                        'error': str(e)
                    })
            
            # Prepare response
            response_data = {
                'success': True,
                'employees_created': employees_created,
                'employees_updated': employees_updated,
                'total_processed': employees_created + employees_updated,
                'errors': errors
            }
            
            if errors:
                response_data['message'] = f'Processed {employees_created + employees_updated} employees with {len(errors)} errors'
            else:
                response_data['message'] = f'Successfully processed {employees_created + employees_updated} employees'
            
            return Response(response_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'error': f'File processing failed: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['get'])
    def download_employee_template(self, request, pk=None):
        """Download CSV template for employee bulk import"""
        template_data = {
            'employee_id': ['EMP001', 'EMP002'],
            'first_name': ['John', 'Jane'],
            'last_name': ['Smith', 'Doe'],
            'middle_initial': ['A', 'M'],
            'email': ['john.smith@company.com', 'jane.doe@company.com'],
            'ssn': ['123-45-6789', '987-65-4321'],
            'date_of_birth': ['1985-03-15', '1990-07-22'],
            'gender': ['M', 'F'],
            'marital_status': ['married', 'single'],
            'phone': ['555-123-4567', '555-987-6543'],
            'address_line1': ['123 Main St', '456 Oak Ave'],
            'address_line2': ['Apt 1', ''],
            'city': ['Boston', 'Cambridge'],
            'state': ['MA', 'MA'],
            'zip_code': ['02101', '02139'],
            'hire_date': ['2020-01-15', '2021-06-01'],
            'job_title': ['Software Engineer', 'Product Manager'],
            'department': ['Engineering', 'Product'],
            'salary': [95000, 105000],
            'hours_per_week': [40, 40],
            'employment_status': ['active', 'active'],
            'medical_coverage_tier': ['family', 'employee_only'],
            'dental_coverage_tier': ['family', 'employee_only'],
            'vision_coverage_tier': ['family', 'employee_only']
        }
        
        df = pd.DataFrame(template_data)
        
        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="employee_import_template.csv"'
        df.to_csv(response, index=False)
        
        return response
    
    @action(detail=True, methods=['get'])
    def download_latest_export(self, request, pk=None):
        """Download the latest completed export for this employer"""
        employer = self.get_object()
        
        # Get the latest completed export job for this employer
        latest_export = ExportJob.objects.filter(
            employer=employer,
            status='completed'
        ).order_by('-created_at').first()
        
        # Always generate a current CSV export regardless of previous export jobs
        
        # Generate CSV export of all employees for this employer
        employees = Employee.objects.filter(employer=employer).select_related('employer')
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{employer.name}_employees_export.csv"'
        
        import csv
        writer = csv.writer(response)
        
        # Write headers
        writer.writerow([
            'Employee ID', 'First Name', 'Last Name', 'Email', 'Phone',
            'Date of Birth', 'SSN', 'Address', 'City', 'State', 'Zip',
            'Hire Date', 'Job Title', 'Department', 'Salary', 'Hours per Week',
            'Employment Status'
        ])
        
        # Write employee data
        for emp in employees:
            writer.writerow([
                emp.employee_id,
                emp.first_name,
                emp.last_name,
                emp.email,
                emp.phone,
                emp.date_of_birth,
                emp.ssn,
                emp.address_line1,
                emp.city,
                emp.state,
                emp.zip_code,
                emp.hire_date,
                emp.job_title,
                emp.department,
                emp.salary,
                emp.hours_per_week,
                emp.employment_status
            ])
        
        return response

class EmployerOfferingViewSet(viewsets.ModelViewSet):
    queryset = EmployerOffering.objects.all()
    serializer_class = EmployerOfferingSerializer
    
    @action(detail=False, methods=['get'])
    def by_employer(self, request):
        employer_id = request.query_params.get('employer_id')
        if employer_id:
            offerings = EmployerOffering.objects.filter(employer_id=employer_id)
            serializer = self.get_serializer(offerings, many=True)
            return Response(serializer.data)
        return Response({'error': 'employer_id parameter required'}, status=400)

class ExportJobViewSet(viewsets.ModelViewSet):
    queryset = ExportJob.objects.all()
    serializer_class = ExportJobSerializer
    
    @action(detail=False, methods=['post'])
    def generate_aetna_export(self, request):
        """Generate Aetna Excel export for an employer"""
        employer_id = request.data.get('employer_id')
        coverage_type = request.data.get('coverage_type', 'medical')
        
        if not employer_id:
            return Response({'error': 'employer_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            employer = Employer.objects.get(id=employer_id)
            aetna_carrier = Carrier.objects.get(name='Aetna')
            
            # Create export job
            export_job = ExportJob.objects.create(
                employer=employer,
                carrier=aetna_carrier,
                coverage_type=coverage_type,
                created_by_id=1,  # For now, using default user
                status='processing'
            )
            
            # Generate the Excel file
            try:
                file_path = self._generate_aetna_excel(employer, coverage_type, export_job)
                export_job.file_name = os.path.basename(file_path)
                export_job.status = 'completed'
                export_job.save()
                
                return Response({
                    'message': 'Aetna export generated successfully',
                    'job_id': str(export_job.id),
                    'status': 'completed',
                    'file_name': export_job.file_name
                })
                
            except Exception as e:
                export_job.status = 'failed'
                export_job.error_details = {'error': str(e)}
                export_job.save()
                return Response({
                    'error': f'Export generation failed: {str(e)}'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
        except Employer.DoesNotExist:
            return Response({'error': 'Employer not found'}, status=status.HTTP_404_NOT_FOUND)
        except Carrier.DoesNotExist:
            return Response({'error': 'Aetna carrier not found'}, status=status.HTTP_404_NOT_FOUND)
    
    def _generate_aetna_excel(self, employer, coverage_type, export_job):
        """Generate Aetna-specific Excel file"""
        # Get employees and their dependents
        employees = Employee.objects.filter(employer=employer).prefetch_related('dependents')
        
        # Prepare data for Excel export
        export_data = []
        
        for employee in employees:
            # Employee row
            employee_data = {
                'Employee ID': employee.employee_id,
                'First Name': employee.first_name,
                'Last Name': employee.last_name,
                'Middle Initial': employee.middle_initial,
                'SSN': employee.ssn,
                'Date of Birth': employee.date_of_birth.strftime('%m/%d/%Y') if employee.date_of_birth else '',
                'Gender': employee.gender,
                'Marital Status': employee.marital_status.title(),
                'Email': employee.email,
                'Phone': employee.phone,
                'Address Line 1': employee.address_line1,
                'Address Line 2': employee.address_line2,
                'City': employee.city,
                'State': employee.state,
                'ZIP Code': employee.zip_code,
                'Hire Date': employee.hire_date.strftime('%m/%d/%Y') if employee.hire_date else '',
                'Job Title': employee.job_title,
                'Department': employee.department,
                'Annual Salary': float(employee.salary),
                'Hours Per Week': float(employee.hours_per_week),
                'Employment Status': employee.employment_status.title(),
                'Coverage Tier': getattr(employee, f'{coverage_type}_coverage_tier', ''),
                'Relationship': 'Employee',
                'Dependent First Name': '',
                'Dependent Last Name': '',
                'Dependent DOB': '',
                'Dependent Gender': '',
                'Dependent SSN': ''
            }
            export_data.append(employee_data)
            
            # Dependent rows
            for dependent in employee.dependents.all():
                dependent_data = employee_data.copy()
                dependent_data.update({
                    'Relationship': dependent.relationship.title(),
                    'Dependent First Name': dependent.first_name,
                    'Dependent Last Name': dependent.last_name,
                    'Dependent DOB': dependent.date_of_birth.strftime('%m/%d/%Y') if dependent.date_of_birth else '',
                    'Dependent Gender': dependent.gender,
                    'Dependent SSN': dependent.ssn
                })
                
                # Only include if dependent has coverage for this type
                coverage_field = f'{coverage_type}_coverage'
                if hasattr(dependent, coverage_field) and getattr(dependent, coverage_field):
                    export_data.append(dependent_data)
        
        # Create DataFrame and Excel file
        df = pd.DataFrame(export_data)
        
        # Create exports directory if it doesn't exist
        exports_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        
        # Generate filename
        filename = f"Aetna_{employer.name.replace(' ', '_')}_{coverage_type}_{date.today().strftime('%Y%m%d')}_{str(uuid.uuid4())[:8]}.xlsx"
        file_path = os.path.join(exports_dir, filename)
        
        # Write to Excel with Aetna-specific formatting
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Census Data', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Census Data']
            
            # Format headers
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        return file_path
    
    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        """Download completed export file"""
        export_job = self.get_object()
        if export_job.status == 'completed' and export_job.file_name:
            file_path = os.path.join(settings.MEDIA_ROOT, 'exports', export_job.file_name)
            
            if os.path.exists(file_path):
                with open(file_path, 'rb') as f:
                    response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = f'attachment; filename="{export_job.file_name}"'
                    return response
            else:
                return Response({'error': 'File not found'}, status=status.HTTP_404_NOT_FOUND)
        
        return Response({'error': 'Export not ready'}, status=status.HTTP_400_BAD_REQUEST)

class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return EmployeeDetailSerializer
        return EmployeeSerializer
    
    @action(detail=False, methods=['get'])
    def by_employer(self, request):
        employer_id = request.query_params.get('employer_id')
        if employer_id:
            employees = Employee.objects.filter(employer_id=employer_id)
            serializer = self.get_serializer(employees, many=True)
            return Response(serializer.data)
        return Response({'error': 'employer_id parameter required'}, status=400)

class DependentViewSet(viewsets.ModelViewSet):
    queryset = Dependent.objects.all()
    serializer_class = DependentSerializer
    
    @action(detail=False, methods=['get'])
    def by_employee(self, request):
        employee_id = request.query_params.get('employee_id')
        if employee_id:
            dependents = Dependent.objects.filter(employee_id=employee_id)
            serializer = self.get_serializer(dependents, many=True)
            return Response(serializer.data)
        return Response({'error': 'employee_id parameter required'}, status=400)

class EnrollmentPeriodViewSet(viewsets.ModelViewSet):
    queryset = EnrollmentPeriod.objects.all()
    serializer_class = EnrollmentPeriodSerializer
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return EnrollmentPeriodDetailSerializer
        return EnrollmentPeriodSerializer
    
    @action(detail=False, methods=['get'])
    def by_employer(self, request):
        employer_id = request.query_params.get('employer_id')
        if employer_id:
            periods = EnrollmentPeriod.objects.filter(employer_id=employer_id)
            serializer = self.get_serializer(periods, many=True)
            return Response(serializer.data)
        return Response({'error': 'employer_id parameter required'}, status=400)
    
    @action(detail=False, methods=['get'])
    def active(self, request):
        """Get active enrollment periods"""
        active_periods = EnrollmentPeriod.objects.filter(
            status='active',
            start_date__lte=date.today(),
            end_date__gte=date.today()
        )
        serializer = self.get_serializer(active_periods, many=True)
        return Response(serializer.data)

class EmployeeEnrollmentViewSet(viewsets.ModelViewSet):
    queryset = EmployeeEnrollment.objects.all()
    serializer_class = EmployeeEnrollmentSerializer
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return EmployeeEnrollmentDetailSerializer
        elif self.action == 'summary':
            return EmployeeEnrollmentSummarySerializer
        return EmployeeEnrollmentSerializer
    
    @action(detail=False, methods=['get'])
    def by_employee(self, request):
        employee_id = request.query_params.get('employee_id')
        if employee_id:
            enrollments = EmployeeEnrollment.objects.filter(employee_id=employee_id)
            serializer = self.get_serializer(enrollments, many=True)
            return Response(serializer.data)
        return Response({'error': 'employee_id parameter required'}, status=400)
    
    @action(detail=False, methods=['get'])
    def by_period(self, request):
        period_id = request.query_params.get('period_id')
        if period_id:
            enrollments = EmployeeEnrollment.objects.filter(enrollment_period_id=period_id)
            serializer = self.get_serializer(enrollments, many=True)
            return Response(serializer.data)
        return Response({'error': 'period_id parameter required'}, status=400)
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get enrollment summary by period"""
        period_id = request.query_params.get('period_id')
        if period_id:
            enrollments = EmployeeEnrollment.objects.filter(enrollment_period_id=period_id)
            serializer = EmployeeEnrollmentSummarySerializer(enrollments, many=True)
            return Response(serializer.data)
        return Response({'error': 'period_id parameter required'}, status=400)
    
    @action(detail=True, methods=['post'])
    def start_enrollment(self, request, pk=None):
        """Start an employee's enrollment process"""
        enrollment = self.get_object()
        if enrollment.status == 'not_started':
            enrollment.status = 'in_progress'
            enrollment.started_at = datetime.now()
            enrollment.save()
            serializer = self.get_serializer(enrollment)
            return Response(serializer.data)
        return Response({'error': 'Enrollment already started'}, status=400)
    
    @action(detail=True, methods=['post'])
    def submit_enrollment(self, request, pk=None):
        """Submit an employee's enrollment for approval"""
        enrollment = self.get_object()
        if enrollment.status == 'in_progress':
            enrollment.status = 'submitted'
            enrollment.submitted_at = datetime.now()
            enrollment.save()
            
            # Create enrollment event
            EnrollmentEvent.objects.create(
                employee=enrollment.employee,
                event_type='enrollment',
                effective_date=enrollment.enrollment_period.coverage_effective_date,
                reason=f'Enrollment submitted for period: {enrollment.enrollment_period.name}'
            )
            
            serializer = self.get_serializer(enrollment)
            return Response(serializer.data)
        return Response({'error': 'Enrollment cannot be submitted'}, status=400)
    
    @action(detail=True, methods=['post'])
    def approve_enrollment(self, request, pk=None):
        """Approve an employee's enrollment"""
        enrollment = self.get_object()
        if enrollment.status == 'submitted':
            enrollment.status = 'approved'
            enrollment.approved_at = datetime.now()
            enrollment.approved_by = request.user
            enrollment.save()
            serializer = self.get_serializer(enrollment)
            return Response(serializer.data)
        return Response({'error': 'Enrollment cannot be approved'}, status=400)

class PlanEnrollmentViewSet(viewsets.ModelViewSet):
    queryset = PlanEnrollment.objects.all()
    serializer_class = PlanEnrollmentSerializer
    
    @action(detail=False, methods=['get'])
    def by_employee_enrollment(self, request):
        enrollment_id = request.query_params.get('enrollment_id')
        if enrollment_id:
            plan_enrollments = PlanEnrollment.objects.filter(employee_enrollment_id=enrollment_id)
            serializer = self.get_serializer(plan_enrollments, many=True)
            return Response(serializer.data)
        return Response({'error': 'enrollment_id parameter required'}, status=400)
    
    @action(detail=False, methods=['get'])
    def by_plan(self, request):
        plan_id = request.query_params.get('plan_id')
        if plan_id:
            plan_enrollments = PlanEnrollment.objects.filter(plan_id=plan_id, status='enrolled')
            serializer = self.get_serializer(plan_enrollments, many=True)
            return Response(serializer.data)
        return Response({'error': 'plan_id parameter required'}, status=400)
    
    @action(detail=True, methods=['post'])
    def terminate(self, request, pk=None):
        """Terminate a plan enrollment"""
        plan_enrollment = self.get_object()
        termination_date = request.data.get('termination_date')
        reason = request.data.get('reason', 'Manual termination')
        
        if not termination_date:
            return Response({'error': 'termination_date is required'}, status=400)
        
        plan_enrollment.status = 'terminated'
        plan_enrollment.termination_date = termination_date
        plan_enrollment.save()
        
        # Create enrollment event
        EnrollmentEvent.objects.create(
            employee=plan_enrollment.employee_enrollment.employee,
            event_type='termination',
            effective_date=termination_date,
            plan_enrollment=plan_enrollment,
            reason=reason,
            processed_by=request.user
        )
        
        serializer = self.get_serializer(plan_enrollment)
        return Response(serializer.data)

class EnrollmentEventViewSet(viewsets.ModelViewSet):
    queryset = EnrollmentEvent.objects.all()
    serializer_class = EnrollmentEventSerializer
    
    @action(detail=False, methods=['get'])
    def by_employee(self, request):
        employee_id = request.query_params.get('employee_id')
        if employee_id:
            events = EnrollmentEvent.objects.filter(employee_id=employee_id)
            serializer = self.get_serializer(events, many=True)
            return Response(serializer.data)
        return Response({'error': 'employee_id parameter required'}, status=400)
    
    @action(detail=False, methods=['get'])
    def recent(self, request):
        """Get recent enrollment events (last 30 days)"""
        from datetime import timedelta
        thirty_days_ago = date.today() - timedelta(days=30)
        events = EnrollmentEvent.objects.filter(effective_date__gte=thirty_days_ago)
        serializer = self.get_serializer(events, many=True)
        return Response(serializer.data)

class EmployeeFormSubmissionViewSet(viewsets.ModelViewSet):
    queryset = EmployeeFormSubmission.objects.all()
    serializer_class = EmployeeFormSubmissionSerializer
    
    def get_serializer_class(self):
        if self.action == 'list':
            return EmployeeFormSubmissionListSerializer
        return EmployeeFormSubmissionSerializer
    
    @action(detail=False, methods=['get'])
    def by_employer(self, request):
        employer_id = request.query_params.get('employer_id')
        if employer_id:
            submissions = EmployeeFormSubmission.objects.filter(employer_id=employer_id)
            serializer = self.get_serializer(submissions, many=True)
            return Response(serializer.data)
        return Response({'error': 'employer_id parameter required'}, status=400)
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve a form submission and create employee record"""
        from django.utils import timezone
        submission = self.get_object()
        
        if submission.status != 'pending':
            return Response({'error': 'Only pending submissions can be approved'}, status=400)
        
        try:
            # Create employee record from form submission
            employee = Employee.objects.create(
                employer=submission.employer,
                employee_id=f"EMP{submission.id.hex[:8].upper()}",
                first_name=submission.first_name,
                last_name=submission.last_name,
                email=submission.email,
                phone=submission.phone,
                date_of_birth=submission.date_of_birth,
                ssn=submission.ssn,
                address_line1=submission.address_line1,
                address_line2=submission.address_line2,
                city=submission.city,
                state=submission.state,
                zip_code=submission.zip_code,
                hire_date=submission.hire_date,
                job_title=submission.job_title,
                department=submission.department,
                salary=submission.salary,
                hours_per_week=submission.hours_per_week,
                employment_status='active',
                gender='M',  # Default, can be updated later
                marital_status='single'  # Default, can be updated later
            )
            
            # Update submission status
            submission.status = 'approved'
            submission.reviewed_by = request.user if request.user.is_authenticated else None
            submission.reviewed_at = timezone.now()
            submission.created_employee = employee
            submission.notes = request.data.get('notes', '')
            submission.save()
            
            return Response({
                'message': 'Form submission approved and employee created',
                'employee_id': employee.id,
                'submission_id': submission.id
            })
            
        except Exception as e:
            return Response({'error': f'Failed to create employee: {str(e)}'}, status=500)
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject a form submission"""
        from django.utils import timezone
        submission = self.get_object()
        
        if submission.status != 'pending':
            return Response({'error': 'Only pending submissions can be rejected'}, status=400)
        
        submission.status = 'rejected'
        submission.reviewed_by = request.user if request.user.is_authenticated else None
        submission.reviewed_at = timezone.now()
        submission.notes = request.data.get('notes', '')
        submission.save()
        
        return Response({'message': 'Form submission rejected'})
    
    @action(detail=True, methods=['post'])
    def request_changes(self, request, pk=None):
        """Request changes to a form submission"""
        from django.utils import timezone
        submission = self.get_object()
        
        if submission.status != 'pending':
            return Response({'error': 'Only pending submissions can have changes requested'}, status=400)
        
        submission.status = 'changes_requested'
        submission.reviewed_by = request.user if request.user.is_authenticated else None
        submission.reviewed_at = timezone.now()
        submission.notes = request.data.get('notes', '')
        submission.save()
        
        return Response({'message': 'Changes requested for form submission'})

class EmployeePortalViewSet(viewsets.ModelViewSet):
    queryset = EmployeePortalUser.objects.all()
    serializer_class = EmployeePortalUserSerializer
    authentication_classes = []  # Bypass DRF auth for custom employee portal auth
    permission_classes = []
    
    @action(detail=False, methods=['post'])
    def register(self, request):
        """Register a new employee portal user"""
        serializer = EmployeePortalRegisterSerializer(data=request.data)
        if serializer.is_valid():
            email = serializer.validated_data['email']
            password = serializer.validated_data['password']
            form_submission_id = serializer.validated_data.get('form_submission_id')
            
            # Check if user already exists
            if EmployeePortalUser.objects.filter(email=email).exists():
                return Response({'error': 'User with this email already exists'}, status=400)
            
            # Create portal user
            portal_user = EmployeePortalUser.objects.create(
                email=email
            )
            portal_user.set_password(password)
            
            # Link to form submission if provided
            if form_submission_id:
                try:
                    form_submission = EmployeeFormSubmission.objects.get(id=form_submission_id, email=email)
                    portal_user.form_submission = form_submission
                except EmployeeFormSubmission.DoesNotExist:
                    return Response({'error': 'Form submission not found or email mismatch'}, status=400)
            
            portal_user.save()
            
            # Generate verification token
            import secrets
            portal_user.email_verification_token = secrets.token_urlsafe(32)
            portal_user.save()
            
            return Response({
                'message': 'Registration successful',
                'user_id': portal_user.id,
                'email': portal_user.email,
                'verification_required': not portal_user.email_verified
            })
        
        return Response(serializer.errors, status=400)
    
    @action(detail=False, methods=['post'])
    def login(self, request):
        """Employee portal login"""
        serializer = EmployeePortalLoginSerializer(data=request.data)
        if serializer.is_valid():
            email = serializer.validated_data['email']
            password = serializer.validated_data['password']
            
            try:
                portal_user = EmployeePortalUser.objects.get(email=email)
                if not portal_user.is_active:
                    return Response({'error': 'Account is disabled'}, status=400)
                
                if portal_user.check_password(password):
                    # Update last login
                    from django.utils import timezone
                    portal_user.last_login = timezone.now()
                    portal_user.save()
                    
                    # Generate session token (simple implementation)
                    import secrets, jwt
                    from django.conf import settings
                    
                    token_payload = {
                        'user_id': str(portal_user.id),
                        'email': portal_user.email,
                        'exp': timezone.now().timestamp() + 86400  # 24 hours
                    }
                    
                    token = jwt.encode(token_payload, getattr(settings, 'SECRET_KEY', 'fallback-secret'), algorithm='HS256')
                    
                    user_serializer = EmployeePortalUserSerializer(portal_user)
                    return Response({
                        'message': 'Login successful',
                        'token': token,
                        'user': user_serializer.data
                    })
                else:
                    return Response({'error': 'Invalid credentials'}, status=400)
            
            except EmployeePortalUser.DoesNotExist:
                return Response({'error': 'Invalid credentials'}, status=400)
        
        return Response(serializer.errors, status=400)
    
    @action(detail=False, methods=['get'])
    def me(self, request):
        """Get current user profile"""
        # Simple token authentication check
        auth_header = request.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Bearer '):
            return Response({'error': 'Authentication required'}, status=401)
        
        token = auth_header.split(' ')[1]
        try:
            import jwt
            from django.conf import settings
            
            payload = jwt.decode(token, getattr(settings, 'SECRET_KEY', 'fallback-secret'), algorithms=['HS256'])
            user_id = payload['user_id']
            
            portal_user = EmployeePortalUser.objects.get(id=user_id)
            serializer = EmployeePortalUserSerializer(portal_user)
            return Response(serializer.data)
            
        except (jwt.InvalidTokenError, EmployeePortalUser.DoesNotExist):
            return Response({'error': 'Invalid token'}, status=401)
    
    @action(detail=True, methods=['post'])
    def verify_email(self, request, pk=None):
        """Verify email with token"""
        portal_user = self.get_object()
        token = request.data.get('token')
        
        if portal_user.email_verification_token == token:
            portal_user.email_verified = True
            portal_user.email_verification_token = ''
            portal_user.save()
            return Response({'message': 'Email verified successfully'})
        
        return Response({'error': 'Invalid verification token'}, status=400)
    
    @action(detail=True, methods=['post'])
    def change_password(self, request, pk=None):
        """Change user password"""
        portal_user = self.get_object()
        current_password = request.data.get('current_password')
        new_password = request.data.get('new_password')
        
        if not portal_user.check_password(current_password):
            return Response({'error': 'Current password is incorrect'}, status=400)
        
        if len(new_password) < 8:
            return Response({'error': 'New password must be at least 8 characters'}, status=400)
        
        portal_user.set_password(new_password)
        portal_user.save()
        
        return Response({'message': 'Password changed successfully'})
